# convert_excel_to_datajs.py
# Usage:
#   python convert_excel_to_datajs.py "FC Shards Calculator.xlsx" > data.js
#
# Liest das Sheet "Skills" aus deiner Excel und erzeugt window.FC_DATA = {...}
import sys, json
import openpyxl

xlsx = sys.argv[1]
wb = openpyxl.load_workbook(xlsx, data_only=True)
sh = wb["Skills"]

headers = [sh.cell(1, c).value for c in range(1, sh.max_column+1)]

pairs = []
for idx, h in enumerate(headers, start=1):
    if isinstance(h, str) and "(fc shards)" in h:
        skill = h.split(" (fc shards)")[0].strip()
        pairs.append((skill, idx, idx+1))

out = {}
for skill, fc_col, iron_col in pairs:
    fc, iron = [], []
    for r in range(2, 14):  # Lv1..Lv12
        fc_val = sh.cell(r, fc_col).value
        iron_val = sh.cell(r, iron_col).value
        fc.append(int(fc_val) if fc_val is not None else 0)
        iron.append(int(iron_val) if iron_val is not None else 0)
    out[skill] = {"fc": fc, "iron": iron}

print("window.FC_DATA = " + json.dumps(out, ensure_ascii=False, indent=2) + ";")
